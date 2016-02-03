﻿__author__="Sully Cothran, Jaimiey Sears"
__copyright__="October 26, 2015"
__version__= 1.0

import math, time, pickle
import datetime as dt
from constants import *
from openpyxl import Workbook
from openpyxl.cell import get_column_letter as toLetter

def decode_new(string, anglePhi):
    # resolution of the scan angle TODO: move to constants.py
    resolution = 0.25

    splitData = splitNparts(string, 3)

    #init the result lists
    dists = []
    x = []
    y = []
    z = []
    phis = []
    thetas = []

    # get starting angle from constants
    angleTheta = START_ANGLE

    # decode all 3-letter string values to numeric values
    for point in splitData:
        dists.append(decodeShort(point))

    # translate to XYZ-coords based on angleTheta and anglePhi
    for dist in dists:

        angleThetaRadians = math.radians(angleTheta)
        #find the cartesian coordinates
        xCoord = dist*math.cos(angleThetaRadians)
        yCoord = dist*math.sin(angleThetaRadians)

        yCoord = yCoord*math.cos(anglePhi)
        zCoord = yCoord*(-1*math.sin(anglePhi))

        # calculate offsets
        yOffset, zOffset = offsetConversion(anglePhi)

        yCoord += yOffset
        zCoord += zOffset

        x.append(xCoord)
        y.append(yCoord)
        z.append(zCoord)
        phis.append(anglePhi)
        thetas.append(angleTheta)

        angleTheta += resolution

    return x, y, z, dists, phis, thetas

def offsetConversion(angle):
    height = Z_OFFSET
    width = Y_OFFSET

    if width == 0:
        startAngle = math.pi/2
    else:
        startAngle = math.atan(height/width)

    dist = math.sqrt(height*height + width*width)
    x = dist*math.cos(startAngle-angle)
    y = dist*math.sin(startAngle-angle)
    return x, y
def splitNparts(string, n):
    doSplit = True
    strList = []
    while(doSplit):
        if string == "":
            return strList
        if(len(string) < n):
            strList.append(string)
            return strList
        else:
            strList.append(string[0:n])
            string = string[n:len(string)]

def decodeShort(dataStr):
    result = 0
    for i in range(len(dataStr)):
        index = len(dataStr)-1-i
        #Get the lower value
        temp = (ord(dataStr[index]) - 0x30) << (i*6)

        #put the value together with rest
        result += temp

    return result

def generateStampedFileName(ext):
    timestamp = dt.datetime.now()
    return "test_vectors_{}_{}_{}_{}_{}_{}{}".format(timestamp.year, timestamp.month, timestamp.day,\
           timestamp.hour, timestamp.minute, timestamp.second, ext)

# Writes an object to a pickle file
def writeToPickle(filename, obj):
    with open(filename, 'wb') as f:
        pickle.dump(obj, f)
        debugPrint("Pickle file {} written".format(filename), ROSTA)


def pickle2xlsx(filename):
    dataArrays = None
    try:
        with open(filename, 'rb') as f:
            dataArrays = pickle.load(f)
    except:
        return OPERATION_FAILURE
    wbSave(generateStampedFileName('.xlsx'), dataArrays)
    debugPrint("Wrote .xlsx file", UTILITY)
    return OPERATION_SUCCESS


##
# debugPrint
#
# Description: prints the specified string only when debug boolean is set to True
#
# __params__
##
def debugPrint(string, lvl):
    if lvl in DEBUG_LEVELS:
        print "{}\t\t{}".format("DEBUG LOG:",string)
    return

# Saves a workbook (MS Excel)
def wbSave(filename, dataArrays):
    # write to excel workbook
    wb = Workbook()
    # outfile = load_workbook(filename=generateStampedFileName(), read_only=False, keep_vba=True)
    sheet1 = wb.active
    # sheet1 = outfile.active
    sheet1['A1'] = "X"
    sheet1['B1'] = "Y"
    sheet1['C1'] = "Z"
    sheet1['D1'] = "Dist"
    sheet1['E1'] = "Phi"
    sheet1['F1'] = "Theta"

    # insert x y z into excel document
    for i in range(len(dataArrays)):
        dataset = dataArrays[i]
        for j in range(len(dataset)):
            cell = '{}{}'.format(toLetter(i+1),j+2)
            sheet1[cell] = dataset[j]

    wb.save(filename)
    debugPrint("workbook saved as {}".format(filename), ROSTA)


## UNIT TESTS FOR DECODE ##
# debugPrint("Unit test 1", UTILITY)
# X, Y, Z, scanAngle, dataOutput, TH, PHI, DIST= decode(str('0CB0'), 0, 0)
# if dataOutput == str('Angle: 0.0 Dist: 123 X: 1234.0 Y: 0 Z: 0\n'): debugPrint( "Long Decode Passed.\n", UTILITY)
# else: debugPrint("Long decode failed with {}\n".format(dataOutput), UTILITY)

debugPrint( "Unit test 2", UTILITY)
result = decodeShort(str('CB'))
if str(result) == str('1234'): debugPrint("Short decode Passed.\n", UTILITY)
else: debugPrint( "Short decode failed with {}".format(result), UTILITY)

debugPrint( "Unit test 3", UTILITY)
result = splitNparts("HelloHelloHello",5)
if result == ["Hello", "Hello","Hello"]: debugPrint( "Split 5 Parts Passed.\n", UTILITY)
else: debugPrint( "Split 5 Failed with {}".format(result), UTILITY)

debugPrint( "Unit test 4", UTILITY)
result = splitNparts("HelloHelloHello",4)
if result == ["Hell", "oHel","loHe","llo"]: debugPrint("Split 4 Parts Passed.\n", UTILITY)
else: debugPrint( "Split 4 Failed with {}".format(result), UTILITY)

# pickle2xlsx("test_vectors_2015_12_10_19_20_21.dat")